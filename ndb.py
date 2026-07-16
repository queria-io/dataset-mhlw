"""NDB オープンデータ 特定健診（検査値）の取得・整形。

厚生労働省 NDB（ナショナルデータベース）オープンデータ 第 11 回（2023 年度）の
「特定健診 検査」ZIP を取得し、検査項目ごとの「都道府県別性年齢階級別分布」Excel を
縦持ち（long 形式）の NDJSON に統合する。

元データは検査項目（BMI・HbA1C・収縮期血圧・眼底検査分類など）ごとに 1 ファイルで、
各ファイルは 都道府県 × 検査値階層（数値区分・分類・所見など）を行に、
性（男／女）× 年齢階級（40〜74 歳を 5 歳刻み）を列に持つクロス集計表になっている。
本モジュールは全ファイルを 1 行 = 1 セルの long 形式に展開する。
型変換・都道府県コード付与は dbt（stg / mart）で行うため、値の数値化のみここで行う。
"""

import io
import json
import re
import urllib.request
import zipfile
from pathlib import Path

import openpyxl

# 第 11 回 NDB オープンデータ（2023 年度）特定健診 検査 ZIP。
# 出典ページ: https://www.mhlw.go.jp/stf/seisakunitsuite/bunya/0000177182.html
KENSHIN_ZIP_URL = "https://www.mhlw.go.jp/content/12400000/001711940.zip"

# 取り込むのは「都道府県別性年齢階級別分布」ファイルのみ。二次医療圏別・
# 「詳細情報レコード含む」・「各項目の平均値」（分布ではなく平均値で構造が異なる）は
# 将来拡張に回す。
_TARGET_SUFFIX = "都道府県別性年齢階級別分布.xlsx"
_EXCLUDE_DIR = "詳細情報レコード含む"
_EXCLUDE_ITEM = "各項目の平均値"

# 列レイアウト（1 始まり）。A=都道府県名 / B=検査値階層 /
# 男 40〜74 歳 = 3〜9・男中計 = 10 / 女 40〜74 歳 = 11〜17・女中計 = 18。
# 中計（小計）は集計値のため取り込まない。
_AGE_CLASSES = ["40～44歳", "45～49歳", "50～54歳", "55～59歳", "60～64歳", "65～69歳", "70～74歳"]
_MALE_COLS = list(range(3, 10))  # 3..9
_FEMALE_COLS = list(range(11, 18))  # 11..17
_DATA_START_ROW = 6

# 全国地方公共団体コード（都道府県コード 2 桁）。特定健診集計は都道府県名で提供されるため
# 名称からコードを引く。「都道府県判別不可」など該当なしは NULL。
_PREFECTURE_CODES = {
    "北海道": "01", "青森県": "02", "岩手県": "03", "宮城県": "04", "秋田県": "05",
    "山形県": "06", "福島県": "07", "茨城県": "08", "栃木県": "09", "群馬県": "10",
    "埼玉県": "11", "千葉県": "12", "東京都": "13", "神奈川県": "14", "新潟県": "15",
    "富山県": "16", "石川県": "17", "福井県": "18", "山梨県": "19", "長野県": "20",
    "岐阜県": "21", "静岡県": "22", "愛知県": "23", "三重県": "24", "滋賀県": "25",
    "京都府": "26", "大阪府": "27", "兵庫県": "28", "奈良県": "29", "和歌山県": "30",
    "鳥取県": "31", "島根県": "32", "岡山県": "33", "広島県": "34", "山口県": "35",
    "徳島県": "36", "香川県": "37", "愛媛県": "38", "高知県": "39", "福岡県": "40",
    "佐賀県": "41", "長崎県": "42", "熊本県": "43", "大分県": "44", "宮崎県": "45",
    "鹿児島県": "46", "沖縄県": "47",
}


def _fiscal_year(title: str | None) -> int | None:
    """タイトル行（例「特定健診（BMI)：2023年度…」）から年度を取り出す。"""
    if not title:
        return None
    m = re.search(r"(\d{4})\s*年度", title)
    return int(m.group(1)) if m else None


def _unit(header_cell: str | None) -> str | None:
    """検査値階層ヘッダ（例「検査値階層\\n(kg/㎡)」）から単位・注記部分を取り出す。

    数値区分の項目は改行後に単位が付く。眼底検査・尿糖・心電図など分類項目は付かない。
    """
    if not header_cell or "\n" not in header_cell:
        return None
    note = header_cell.split("\n", 1)[1].strip()
    return note or None


def _to_count(value: object) -> int | None:
    """値セルを人数（整数）へ。集計値 10 未満のマスク（「‐」等）や空欄は NULL。"""
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return int(value)
    return None


def _flatten_sheet(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[dict]:
    """1 検査項目のシートを long 形式レコードへ展開する。"""
    test_item = ws.title.strip()
    fiscal_year = _fiscal_year(ws.cell(1, 1).value)
    unit = _unit(ws.cell(2, 2).value)

    records: list[dict] = []
    prefecture: str | None = None
    for row in range(_DATA_START_ROW, ws.max_row + 1):
        pref_cell = ws.cell(row, 1).value
        if pref_cell is not None and str(pref_cell).strip():
            prefecture = str(pref_cell).strip()
        value_class = ws.cell(row, 2).value
        if value_class is None or not str(value_class).strip():
            continue
        value_class = str(value_class).strip()
        prefecture_code = _PREFECTURE_CODES.get(prefecture or "")
        for sex, cols in (("男", _MALE_COLS), ("女", _FEMALE_COLS)):
            for age_class, col in zip(_AGE_CLASSES, cols):
                records.append(
                    {
                        "fiscal_year": fiscal_year,
                        "test_item": test_item,
                        "unit": unit,
                        "prefecture": prefecture,
                        "prefecture_code": prefecture_code,
                        "value_class": value_class,
                        "sex": sex,
                        "age_class": age_class,
                        "count": _to_count(ws.cell(row, col).value),
                    }
                )
    return records


def flatten_zip(zip_bytes: bytes, ndjson_path: Path) -> int:
    """特定健診 検査 ZIP を展開し、対象 Excel を統合 NDJSON へ。行数を返す。"""
    rows = 0
    with (
        zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf,
        ndjson_path.open("w", encoding="utf-8") as out,
    ):
        for name in sorted(zf.namelist()):
            if not name.endswith(_TARGET_SUFFIX):
                continue
            if _EXCLUDE_DIR in name or _EXCLUDE_ITEM in name:
                continue
            wb = openpyxl.load_workbook(io.BytesIO(zf.read(name)), data_only=True)
            for record in _flatten_sheet(wb[wb.sheetnames[0]]):
                out.write(json.dumps(record, ensure_ascii=False) + "\n")
                rows += 1
    return rows


def download_and_flatten(ndjson_path: Path) -> int:
    """特定健診 検査 ZIP を取得し、統合 NDJSON を書き出す。行数を返す。"""
    with urllib.request.urlopen(KENSHIN_ZIP_URL) as resp:
        zip_bytes = resp.read()
    return flatten_zip(zip_bytes, ndjson_path)
